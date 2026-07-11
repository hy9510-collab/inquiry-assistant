# 의정활동 비서 — 웹페이지 생성기
# 사용법:
#   python 대시보드_생성.py                          → 대시보드.html 생성
#   python 대시보드_생성.py --share 의원명 2026-07   → 04_리포트/공유_YYYYMM_의원명.html 생성(인쇄용)
# 데이터 원본: 03_대장/의정활동대장.xlsx (원본 하나, 출력 여러 개)
import json, sys, datetime
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(__file__).parent
SRC = BASE / "03_대장" / "의정활동대장.xlsx"
SHEETS = ["공약이행", "민원처리", "의회활동", "보도홍보", "일정과제"]

def _v(c):
    if c.value is None: return ""
    if isinstance(c.value, (datetime.date, datetime.datetime)): return c.value.strftime("%Y-%m-%d")
    return str(c.value).strip()

def read(src=SRC):
    wb = load_workbook(src, data_only=True)
    data = {}
    for name in SHEETS:
        ws = wb[name]
        headers = [c.value for c in ws[1] if c.value]
        rows = []
        for r in ws.iter_rows(min_row=2, max_col=len(headers)):
            vals = [_v(c) for c in r]
            if any(vals): rows.append(vals)
        data[name] = {"headers": headers, "rows": rows}
    return data

def report_files():
    d = BASE / "04_리포트"
    if not d.exists(): return []
    return sorted([f.name for f in d.iterdir() if f.suffix.lower() in (".hwpx",".docx",".html",".pdf") and not f.name.startswith("~")], reverse=True)

CSS = """
:root{--navy:#1f3864;--bg:#f4f6fa;--line:#dde3ee;--acc:#3d6cb9}
*{box-sizing:border-box}body{font-family:'Malgun Gothic','맑은 고딕',sans-serif;margin:0;background:var(--bg);color:#222}
header{background:var(--navy);color:#fff;padding:14px 24px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap}
header h1{margin:0;font-size:19px}header .sub{font-size:11.5px;opacity:.8}
nav{display:flex;gap:4px;background:#fff;border-bottom:1px solid var(--line);padding:0 20px;overflow-x:auto}
nav .tab{padding:12px 18px;cursor:pointer;font-size:14px;border-bottom:3px solid transparent;white-space:nowrap}
nav .tab.on{border-color:var(--navy);color:var(--navy);font-weight:700}
.wrap{max-width:1250px;margin:0 auto;padding:20px}
.grid2{display:grid;grid-template-columns:repeat(auto-fit,minmax(420px,1fr));gap:16px}
.panel{background:#fff;border:1px solid var(--line);border-radius:12px;padding:18px}
.panel h2{margin:0 0 10px;font-size:16px;color:var(--navy)}
.panel h3{margin:14px 0 6px;font-size:13px;color:#556}
.kpis{display:flex;gap:10px;flex-wrap:wrap;margin:8px 0}
.kpi{flex:1;min-width:90px;background:var(--bg);border-radius:8px;padding:10px;text-align:center}
.kpi .n{font-size:22px;font-weight:700;color:var(--navy)}.kpi .t{font-size:11px;color:#667}
.badge{display:inline-block;padding:2px 9px;border-radius:10px;font-size:11px;background:#eef2fa;color:var(--navy);white-space:nowrap}
.b-완료,.b-회신완료{background:#e2f2e5;color:#1a7a2e}.b-추진중,.b-처리중,.b-실시,.b-일부이행,.b-후속조치중,.b-진행중{background:#fff3d6;color:#96660a}
.b-보류,.b-취소{background:#fde3e3;color:#a12626}.b-예정,.b-구상,.b-접수{background:#e8eefb;color:#2a4d8f}
.ctrl{display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap}
select,input{padding:7px 10px;border:1px solid var(--line);border-radius:6px;font-size:13px;font-family:inherit}
input.q{flex:1;min-width:160px}
table{width:100%;border-collapse:collapse;background:#fff;font-size:12.5px}
th{background:#e8edf7;color:var(--navy);padding:8px;border:1px solid var(--line);position:sticky;top:0}
td{padding:7px 8px;border:1px solid var(--line);vertical-align:top}td a{color:var(--acc)}
.tablebox{overflow:auto;max-height:62vh;border-radius:8px;border:1px solid var(--line)}
.kanban{display:grid;grid-template-columns:repeat(5,1fr);gap:10px}
.kcol{background:#eef1f7;border-radius:10px;padding:10px;min-height:120px}
.kcol h4{margin:0 0 8px;font-size:12.5px;color:var(--navy);text-align:center}
.kcard{background:#fff;border:1px solid var(--line);border-radius:8px;padding:8px;margin-bottom:8px;font-size:12px}
.kcard .m{color:#889;font-size:11px}
.pcard{background:#fff;border:1px solid var(--line);border-radius:10px;padding:14px;margin-bottom:12px}
.pcard .top{display:flex;justify-content:space-between;align-items:center;gap:8px;flex-wrap:wrap}
.pcard .name{font-weight:700;font-size:14px}
.bar{height:6px;background:#e8ecf5;border-radius:3px;margin:8px 0}.bar i{display:block;height:6px;border-radius:3px;background:var(--acc)}
.meta{font-size:12px;color:#556;margin-top:6px;white-space:pre-wrap}
.empty{padding:44px;text-align:center;color:#889;background:#fff;border:1px dashed var(--line);border-radius:10px}
.item{padding:6px 0;border-bottom:1px solid #f0f2f7;font-size:12.5px}.item:last-child{border:none}
.filelist a{display:block;padding:10px 12px;border:1px solid var(--line);border-radius:8px;margin-bottom:8px;background:#fff;color:var(--navy);text-decoration:none;font-size:13px}
footer{font-size:11px;color:#889;padding:16px 24px;text-align:center}
@media(max-width:900px){.kanban{grid-template-columns:1fr 1fr}.grid2{grid-template-columns:1fr}}
"""

JS = r"""
const D=DATA, RPT=REPORTS;
const TABS=["홈","공약","민원","의회활동","보도","일정","리포트"];
let cur="홈";
const esc=s=>String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;");
const fmt=v=>/^https?:\/\//.test(v)?'<a href="'+esc(v)+'" target="_blank">원문</a>':esc(v);
const badge=v=>v?'<span class="badge b-'+esc(v)+'">'+esc(v)+'</span>':"";
const members=[...new Set([].concat(D["공약이행"].rows.map(r=>r[0]),D["민원처리"].rows.map(r=>r[0]),D["의회활동"].rows.map(r=>r[0]),D["보도홍보"].rows.map(r=>r[0])).filter(x=>x))];
const today=new Date().toISOString().slice(0,10);

function nav(){document.getElementById("nav").innerHTML=TABS.map(t=>'<div class="tab'+(t===cur?' on':'')+'" onclick="go(\''+t+'\')">'+t+'</div>').join("")}
function go(t){cur=t;nav();render()}
function tbl(sheet,rows,badgeCol){
  const h=D[sheet].headers;
  if(!rows.length)return '<div class="empty">기록이 없습니다. 03_대장에 입력 후 대시보드를 다시 생성하세요.</div>';
  return '<div class="tablebox"><table><tr>'+h.map(x=>'<th>'+esc(x)+'</th>').join("")+'</tr>'+
    rows.map(r=>'<tr>'+r.map((v,i)=>'<td>'+(i===badgeCol?badge(v):fmt(v))+'</td>').join("")+'</tr>').join("")+'</table></div>';
}
function selFilter(id,rows,col){const e=document.getElementById(id);const v=e?e.value:"";return v?rows.filter(r=>r[col]===v):rows}
function qFilter(rows){const e=document.getElementById("q");const q=(e?e.value:"").toLowerCase();return q?rows.filter(r=>r.join(" ").toLowerCase().includes(q)):rows}
function ctrl(opts){
  return '<div class="ctrl">'+opts.map(o=>'<select id="'+o.id+'" onchange="render()"><option value="">'+o.label+'</option>'+o.values.map(v=>'<option>'+esc(v)+'</option>').join("")+'</select>').join("")+'<input class="q" id="q" placeholder="검색…" oninput="render()"></div>';
}
function keep(ids){const s={};ids.concat(["q"]).forEach(i=>{const e=document.getElementById(i);if(e)s[i]=e.value});return s}
function restore(s){Object.entries(s).forEach(([i,v])=>{const e=document.getElementById(i);if(e)e.value=v})}

function home(){
  if(!members.length&&!D["일정과제"].rows.length)return '<div class="empty">아직 기록이 없습니다. 의원 배정 후 03_대장에 입력하면 여기에 요약이 나타납니다.</div>';
  let h='<div class="grid2">';
  for(const m of members){
    const pl=D["공약이행"].rows.filter(r=>r[0]===m);
    const done=pl.filter(r=>["완료","일부이행"].includes(r[7])).length;
    const mw=D["민원처리"].rows.filter(r=>r[0]===m&&r[6]&&r[6]!=="회신완료").length;
    const acts=D["의회활동"].rows.filter(r=>r[0]===m).sort((a,b)=>b[1].localeCompare(a[1])).slice(0,3);
    h+='<div class="panel"><h2>'+esc(m)+' 의원</h2><div class="kpis">'+
      '<div class="kpi"><div class="n">'+pl.length+'</div><div class="t">공약</div></div>'+
      '<div class="kpi"><div class="n">'+done+'</div><div class="t">완료·일부이행</div></div>'+
      '<div class="kpi"><div class="n">'+mw+'</div><div class="t">미처리 민원</div></div>'+
      '<div class="kpi"><div class="n">'+D["의회활동"].rows.filter(r=>r[0]===m).length+'</div><div class="t">의회활동 누적</div></div></div>'+
      '<h3>최근 활동</h3>'+(acts.length?acts.map(r=>'<div class="item">'+esc(r[1])+' ｜ '+badge(r[2])+' '+esc(r[3])+'</div>').join(""):'<div class="item" style="color:#889">기록 없음</div>')+'</div>';
  }
  const up=D["일정과제"].rows.filter(r=>r[0]>=today&&r[6]!=="완료").sort((a,b)=>a[0].localeCompare(b[0])).slice(0,8);
  h+='<div class="panel"><h2>다가오는 일정</h2>'+(up.length?up.map(r=>'<div class="item">'+esc(r[0])+' ｜ '+badge(r[1])+' '+esc(r[2])+(r[3]?' — '+esc(r[3]):'')+'</div>').join(""):'<div class="item" style="color:#889">예정 일정 없음</div>')+'</div>';
  return h+'</div>';
}
function pledges(){
  let rows=qFilter(selFilter("f_m",selFilter("f_s",D["공약이행"].rows,7),0));
  let h=ctrl([{id:"f_m",label:"전체 의원",values:members},{id:"f_s",label:"전체 상태",values:["구상","추진중","일부이행","완료","보류"]}]);
  if(!rows.length)return h+'<div class="empty">공약 기록이 없습니다. 선거공보 원문 기준으로 03_대장 공약이행 시트에 입력하세요.</div>';
  const pct={"구상":10,"추진중":45,"일부이행":70,"완료":100,"보류":0};
  h+=rows.map(r=>'<div class="pcard"><div class="top"><span class="name">'+esc(r[2])+'</span>'+badge(r[7])+'</div>'+
    '<div class="bar"><i style="width:'+(pct[r[7]]||0)+'%"></i></div>'+
    '<div class="meta">'+esc(r[0])+' ｜ '+esc(r[3])+' ｜ '+esc(r[4])+' ｜ 이행수단: '+esc(r[5])+' ｜ 소관: '+esc(r[6])+'</div>'+
    (r[8]?'<div class="meta"><b>추진내용</b>\n'+esc(r[8])+'</div>':'')+
    (r[9]?'<div class="meta"><b>다음 조치</b> '+esc(r[9])+'</div>':'')+
    (r[10]?'<div class="meta">'+fmt(r[10])+'</div>':'')+'</div>').join("");
  return h;
}
function complaints(){
  const st=["접수","검토","부서이송","처리중","회신완료"];
  let rows=qFilter(selFilter("f_m",D["민원처리"].rows,0));
  let h=ctrl([{id:"f_m",label:"전체 의원",values:members}]);
  if(!rows.length)return h+'<div class="empty">민원 기록이 없습니다.</div>';
  h+='<div class="kanban">'+st.map(s=>'<div class="kcol"><h4>'+s+' ('+rows.filter(r=>r[6]===s).length+')</h4>'+
    rows.filter(r=>r[6]===s).map(r=>'<div class="kcard"><b>'+esc(r[3])+'</b> '+esc(r[2])+'<br>'+esc(r[4]).slice(0,60)+'<div class="m">'+esc(r[0])+' ｜ 접수 '+esc(r[1])+(r[5]?' ｜ '+esc(r[5]):'')+'</div></div>').join("")+'</div>').join("")+'</div>';
  return h;
}
function activities(){
  let rows=qFilter(selFilter("f_m",selFilter("f_t",D["의회활동"].rows,2),0)).sort((a,b)=>b[1].localeCompare(a[1]));
  return ctrl([{id:"f_m",label:"전체 의원",values:members},{id:"f_t",label:"전체 유형",values:["도정질문","5분자유발언","조례발의","상임위질의","행정사무감사","예산심사","토론회·간담회","연구단체","기타"]}])+tbl("의회활동",rows,8);
}
function press(){
  let rows=qFilter(selFilter("f_m",D["보도홍보"].rows,0)).sort((a,b)=>b[1].localeCompare(a[1]));
  return ctrl([{id:"f_m",label:"전체 의원",values:members}])+tbl("보도홍보",rows,-1);
}
function schedule(){
  let rows=qFilter(D["일정과제"].rows).sort((a,b)=>a[0].localeCompare(b[0]));
  return ctrl([])+tbl("일정과제",rows,6);
}
function reports(){
  if(!RPT.length)return '<div class="empty">04_리포트 폴더에 저장된 리포트가 없습니다.</div>';
  return '<div class="filelist">'+RPT.map(f=>'<a href="04_리포트/'+encodeURIComponent(f)+'" target="_blank">📄 '+esc(f)+'</a>').join("")+'</div>';
}
function render(){
  const saved=keep(["f_m","f_s","f_t"]);
  const v={"홈":home,"공약":pledges,"민원":complaints,"의회활동":activities,"보도":press,"일정":schedule,"리포트":reports}[cur]();
  document.getElementById("view").innerHTML=v;
  restore(saved);
}
nav();render();
"""

def build_dashboard(src=SRC, out=None):
    out = out or (BASE / "대시보드.html")
    data = read(src)
    html = f"""<!DOCTYPE html><html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>의정활동 비서</title>
<style>{CSS}</style></head><body>
<header><div><h1>의정활동 비서</h1><div class="sub">경기도의회 문화체육관광위원회 ｜ 정책지원관</div></div>
<div class="sub">생성 {datetime.date.today().isoformat()} ｜ 원본: 03_대장/의정활동대장.xlsx</div></header>
<nav id="nav"></nav>
<div class="wrap" id="view"></div>
<footer>원본 하나, 출력 여러 개 ｜ 모든 기록은 날짜·출처 병기 ｜ 민원인 비식별 처리</footer>
<script>const DATA={json.dumps(data, ensure_ascii=False)};const REPORTS={json.dumps(report_files(), ensure_ascii=False)};{JS}</script>
</body></html>"""
    Path(out).write_text(html, encoding="utf-8")
    print(f"생성 완료: {out}")

def build_share(member, month, src=SRC):
    data = read(src)
    ym = month.replace("-", "")
    def pick(s, dc):
        return [r for r in data[s]["rows"] if r[0] == member and str(r[dc]).startswith(month)]
    pl = [r for r in data["공약이행"]["rows"] if r[0] == member]
    acts, mws, prs = pick("의회활동", 1), pick("민원처리", 1), pick("보도홍보", 1)
    mw_all = [r for r in data["민원처리"]["rows"] if r[0] == member]
    nxt = sorted([r for r in data["일정과제"]["rows"] if r[0][:7] > month], key=lambda r: r[0])[:10]
    stat = {s: sum(1 for r in pl if r[7] == s) for s in ["완료", "일부이행", "추진중", "구상", "보류"]}
    esc = lambda s: str(s).replace("&", "&amp;").replace("<", "&lt;")
    row = lambda cells: "<tr>" + "".join(f"<td>{esc(c)}</td>" for c in cells) + "</tr>"
    def T(hd, rows, empty):
        if not rows: return f'<p class="none">{empty}</p>'
        return "<table><tr>" + "".join(f"<th>{h}</th>" for h in hd) + "</tr>" + "".join(rows) + "</table>"
    pending = sum(1 for r in mw_all if r[6] and r[6] != "회신완료")
    html = f"""<!DOCTYPE html><html lang="ko"><head><meta charset="utf-8"><title>의정활동 리포트 {month} {member}</title>
<style>
body{{font-family:'Malgun Gothic','맑은 고딕',sans-serif;max-width:800px;margin:0 auto;padding:36px 28px;color:#111;font-size:13px;line-height:1.6}}
h1{{text-align:center;color:#1f3864;font-size:22px;margin-bottom:2px}}
.sub{{text-align:center;color:#555;font-size:12px;border-bottom:2px solid #1f3864;padding-bottom:10px;margin-bottom:18px}}
h2{{color:#1f3864;font-size:15px;border-left:4px solid #1f3864;padding-left:8px;margin:22px 0 8px}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th{{background:#e8edf7;color:#1f3864;padding:6px;border:1px solid #ccd4e3}}
td{{padding:6px;border:1px solid #ccd4e3;vertical-align:top}}
.none{{color:#888;font-size:12px}}
.stat{{display:flex;gap:8px;margin:8px 0}}.stat div{{flex:1;text-align:center;background:#f2f5fb;border-radius:6px;padding:8px;font-size:12px}}.stat b{{display:block;font-size:18px;color:#1f3864}}
footer{{margin-top:28px;font-size:11px;color:#777;border-top:1px solid #ddd;padding-top:8px}}
@media print{{body{{padding:0}}.noprint{{display:none}}}}
</style></head><body>
<div class="noprint" style="text-align:right"><button onclick="print()">인쇄 / PDF 저장</button></div>
<h1>의정활동 월간 리포트</h1>
<div class="sub">{esc(member)} 의원 ｜ {month} ｜ 작성: 정책지원관 ｜ 경기도의회 문화체육관광위원회</div>
<h2>1. 이달의 활동 요약</h2>
{T(["구분","건수","비고"],[row(["의회 활동",f"{len(acts)}건",", ".join(sorted(set(r[2] for r in acts))) or "-"]),row(["민원",f"접수 {len(mws)}건",f"누적 미처리 {pending}건"]),row(["언론·홍보",f"{len(prs)}건",", ".join(r[3] for r in prs[:3]) or "-"])],"")}
<h2>2. 공약 이행 현황</h2>
<div class="stat"><div><b>{stat['완료']}</b>완료</div><div><b>{stat['일부이행']}</b>일부이행</div><div><b>{stat['추진중']}</b>추진중</div><div><b>{stat['구상']}</b>구상</div><div><b>{stat['보류']}</b>보류</div></div>
{T(["공약","상태","최근 추진내용 / 다음 조치"],[row([r[2],r[7],(r[8] or "-")+((" / 다음: "+r[9]) if r[9] else "")]) for r in pl],"공약 기록 없음 — 선거공보 기준 입력 예정")}
<h2>3. 이달 민원 처리(비식별)</h2>
{T(["접수일","유형","내용 요지","상태"],[row([r[1],r[3],r[4][:40],r[6]]) for r in mws],"이달 접수 민원 없음")}
<h2>4. 의회 활동 상세</h2>
{T(["일자","유형","제목·안건","후속조치"],[row([r[1],r[2],r[3],r[6] or "-"]) for r in sorted(acts,key=lambda r:r[1])],"이달 의회 활동 기록 없음")}
<h2>5. 다음 일정</h2>
{T(["일자","구분","내용"],[row([r[0],r[1],r[2]]) for r in nxt],"예정 일정 없음")}
<footer>본 리포트는 의정활동대장 원본에서 생성되었습니다. 원문 링크·상세 기록은 대시보드에서 확인 가능합니다. ｜ 생성일 {datetime.date.today().isoformat()}</footer>
</body></html>"""
    out = BASE / "04_리포트" / f"공유_{ym}_{member}.html"
    out.write_text(html, encoding="utf-8")
    print(f"생성 완료: {out}")

if __name__ == "__main__":
    if len(sys.argv) >= 4 and sys.argv[1] == "--share":
        build_share(sys.argv[2], sys.argv[3])
    else:
        build_dashboard()
